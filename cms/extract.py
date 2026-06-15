"""Trích dữ liệu CÓ CẤU TRÚC từ file estimate (.xlsx) theo template MOR:
header opportunity + danh sách chức năng WBS. Dùng cho tính năng 'upload → tự fill'.

Lưu ý: bám theo bố cục template estimate của MOR (sheet Overall/Assumption/WBS Function List).
Trường nào không tìm thấy → để trống (người dùng tự bổ sung).
"""
import os
import re
from openpyxl import load_workbook

NUM_RE = re.compile(r"^\d+(\.\d+)?$")

TECH_KEYWORDS = [
    "React Native", "React", "Node.js", "Express", "Java", "Spring Boot", "MySQL",
    "MongoDB", "Redis", "Stripe", "Socket.io", "Azure", "AWS", "Expo", "BullMQ",
    "Sequelize", "Firebase", "Twilio", "SendGrid", "Vite", "Flutter", "Kotlin",
    "Swift", "PostgreSQL", "Next.js", "Django", "Python", "GraphQL", "Docker",
]
DOMAIN_MAP = [
    ("driver", "Ride-hailing / Đặt xe"), ("ride", "Ride-hailing / Đặt xe"),
    ("insurance", "Bảo hiểm"), ("logistics", "Logistics / WMS"), ("wms", "Logistics / WMS"),
    ("epump", "Bán lẻ xăng dầu / ePump"), ("music", "Sự kiện âm nhạc"),
    ("conference", "Hội nghị / Conference"), ("glueup", "Quản lý sự kiện / cộng đồng"),
]


def extract_casestudy(path):
    """Trích metadata case study từ tài liệu (pptx/pdf/docx...). Trả về dict các trường."""
    from .parsers import extract_blocks
    blocks = extract_blocks(path)
    full = "\n".join(t for _, t in blocks)
    fname = os.path.basename(path).lower()

    first = blocks[0][1] if blocks else ""
    lines = [l for l in first.split("\n") if l.strip()]
    title = lines[1] if len(lines) > 1 else (lines[0] if lines else os.path.basename(path))

    techs = [k for k in TECH_KEYWORDS if re.search(r"\b" + re.escape(k) + r"\b", full, re.I)]
    domain = next((d for key, d in DOMAIN_MAP if key in fname), None)
    customer = "Ẩn theo NDA"
    for brand in ["FWD", "GlueUp", "Glue Up"]:
        if re.search(re.escape(brand), full, re.I):
            customer = brand
            break
    return {
        "name": os.path.splitext(os.path.basename(path))[0],
        "title": title.strip()[:300],
        "customer": customer,
        "domain": domain,
        "tech_stack": ", ".join(dict.fromkeys(techs)) or None,
    }


def _norm(v):
    return str(v).strip() if v is not None else ""


def _num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _find_after(rows, label, max_gap=5):
    for r in rows:
        for ci, v in enumerate(r):
            if _norm(v).startswith(label):
                for nv in r[ci + 1: ci + 1 + max_gap]:
                    if _norm(nv):
                        return _norm(nv)
    return None


def extract_proposal(path):
    """Trả về (header: dict, items: list[dict]). Chỉ áp dụng cho .xlsx template estimate."""
    wb = load_workbook(path, data_only=True)
    overall = list(wb["Overall"].iter_rows(values_only=True)) if "Overall" in wb.sheetnames else []
    assum = list(wb["Assumption"].iter_rows(values_only=True)) if "Assumption" in wb.sheetnames else []
    wbs_name = next((s for s in wb.sheetnames if s.startswith("WBS Function List (MOR)")), None)
    wbs_rows = list(wb[wbs_name].iter_rows(values_only=True)) if wbs_name else []

    tech = _find_after(assum, "Frontend") or _find_after(assum, "Language/ Development")
    if tech:
        tech = tech.replace("\n", ", ").strip()

    header = {
        "name": _find_after(overall, "Project name"),
        "customer": _find_after(overall, "Customer"),
        "department": _find_after(overall, "Department"),
        "doc_type": "estimate",
        "tech_stack": tech,
        "total_effort_mm": _num(_find_after(overall, "Total")),
        "total_effort_md": _num(_find_after(wbs_rows, "Total（MD）")),
        "timeline_months": None,
        "language": _find_after(assum, "Language Support"),
        "source_date": None,
        "status": "draft",
        "description": None,
    }

    items, category = [], None
    for r in wbs_rows:
        c1 = _norm(r[1]) if len(r) > 1 else ""
        if not NUM_RE.match(c1):
            continue
        cat = _norm(r[2]) if len(r) > 2 else ""
        if cat and not NUM_RE.match(cat):
            category = cat
        name = (_norm(r[3]) if len(r) > 3 else "") or (_norm(r[4]) if len(r) > 4 else "") or category
        desc = _norm(r[5]) if len(r) > 5 else ""
        note = _norm(r[12]) if len(r) > 12 else ""
        m = re.search(r"Độ quan trọng[:\s]*([^\n【]+)", note)
        items.append({
            "category": category, "name": name, "description": desc,
            "effort_study": _num(r[6]) if len(r) > 6 else None,
            "effort_fe": _num(r[7]) if len(r) > 7 else None,
            "effort_be": _num(r[8]) if len(r) > 8 else None,
            "effort_ut": _num(r[9]) if len(r) > 9 else None,
            "effort_total": _num(r[10]) if len(r) > 10 else None,
            "priority": m.group(1).strip() if m else None,
        })
    return header, items
